"""
Module: report_utils
Description: This module contains functions to create
and modify a report file.

Author: Richard Saldanha
Date: December 9, 2023
"""

from typing import Any, OrderedDict
import pandas as pd
import openpyxl as px
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ReportOps:  # pylint: disable=R0903 # Need this functionality in separate class for future use
    """
    Contains methods to generate formatted reports.
    """

    def generate_excel_report_data(self, report_path: str, df_data_dict: dict, props: Any) -> dict:
        """
        Accepts a report path, a dictionary object that contains sheet names
        and dataframe values for sheet content, and properties. It converts 
        this data into a formatted excel report that is stored in 
        specified report path as well as returned as a dictionary.

        Args:
            report_path (str): path of the report
            df_data_dict (dict): dictionary object that contains sheet names and
            dataframe values for sheet content
            props (Any): python object containing properties

        Returns:
            dict: dictionary containing report data
        """
        # Write data to Excel file and retrieve report data
        report_data = self.write_to_excel(report_path, df_data_dict)

        # Apply formatting to the Excel file
        self.apply_formatting(report_path, props=props)

        return report_data

    def write_to_excel(self, report_path: str, df_data_dict: dict) -> dict:
        """
        Writes data from DataFrame dictionary to an Excel file.

        Args:
            report_path (str): Path of the report
            df_data_dict (dict): Dictionary containing sheet names and DataFrame values

        Returns:
            dict: Dictionary containing report data
        """
        report_data = {}
        # Open an Excel writer
        with pd.ExcelWriter( # pylint: disable=E0110 # not abstract class
            report_path, engine="openpyxl"
        ) as writer:
            for sheet_name, df_sheet in df_data_dict.items():
                # Merge dataframes if necessary
                df_sheet = self.merge_dataframes(df_sheet)
                # Write the DataFrame to Excel
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                # Store the data for the report
                report_data[sheet_name] = OrderedDict(df_sheet.to_dict(orient="list"))
        return report_data

    def apply_formatting(self, report_path: str, props: Any) -> None:
        """
        Applies formatting to the Excel file.

        Args:
            report_path (str): Path of the report
            df_data_dict (dict): Dictionary containing sheet names and DataFrame values
        """
        # Load the workbook
        workbook = px.load_workbook(report_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Apply formatting to cells
            self.format_cells(sheet, props=props)
            self.equalize_column_lengths(sheet, props=props)
            self.apply_header_styles(sheet, props=props)

        # Save the modified workbook
        workbook.save(report_path)

    def format_cells(self, sheet, props: Any):
        """
        Formats individual cells in the sheet.

        Args:
            sheet: Worksheet object
            props (Any): python object that contains properties from yaml
        """
        cell_border = Border(
            left=Side(style=props.common_report_props.styles.cell.border),
            right=Side(style=props.common_report_props.styles.cell.border),
            top=Side(style=props.common_report_props.styles.cell.border),
            bottom=Side(style=props.common_report_props.styles.cell.border),
        )

        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
        ):
            sheet.row_dimensions[row[0].row].height = props.common_report_props.styles.sheet.row.height
            for cell in row:
                cell.font = Font(size=props.common_report_props.styles.cell.font.size)
                cell.border = cell_border
                try:
                    numeric_value = float(cell.value)
                    if isinstance(numeric_value, (int, float)):
                        cell.alignment = Alignment(horizontal=props.common_report_props.styles.cell.align.center)
                except (ValueError, TypeError):
                    cell.alignment = Alignment(
                        wrap_text=True, shrink_to_fit=True, horizontal=props.common_report_props.styles.cell.align.justify
                    )

    def equalize_column_lengths(self, sheet, props: Any):
        """
        Equalizes column lengths in the sheet.

        Args:
            sheet: Worksheet object
            props (Any) : python object containing properties from yaml
        """
        column_lengths = [
            max(len(str(cell.value)) for cell in column) for column in sheet.iter_cols()
        ]
        for col_idx, _ in enumerate(column_lengths, start=1):
            col_letter = px.utils.get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].hidden = False
            sheet.column_dimensions[col_letter].width = props.common_report_props.styles.sheet.column.width

    def apply_header_styles(self, sheet, props):
        """
        Applies styles to header cells in the sheet.

        Args:
            sheet: Worksheet object
        """
        fg_color = px.styles.colors.Color(rgb=props.common_report_props.styles.header.cell_color)
        header_fill = PatternFill(fgColor=fg_color, patternType="solid")
        header_font = Font(bold=True, size=props.common_report_props.styles.cell.font.size)
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal=props.common_report_props.styles.cell.align.center,
                                       vertical=props.common_report_props.styles.cell.align.center
                                       )
            cell.fill = header_fill
            cell.font = header_font

    def merge_dataframes(self, df_list):
        """
        Merges a list of DataFrames into a single DataFrame.

        Args:
            df_list (list): List of DataFrames

        Returns:
            pd.DataFrame: Merged DataFrame
        """
        return pd.concat(df_list, ignore_index=True)
